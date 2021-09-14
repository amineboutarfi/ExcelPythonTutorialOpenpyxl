from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Seller1": {
		"product1": 78,
		"product2": 62,
		"product3": 14,
		"product4": 8
	},
	"Seller2": {
		"product1": 75,
		"product2": 64,
		"product3": 12,
		"product4": 9
	},
	"Seller3": {
		"product1": 77,
		"product2": 62,
		"product3": 11,
		"product4": 9
	},
	"Seller4": {
		"product1": 75,
		"product2": 63,
		"product3": 14,
		"product4": 9
	},
	"Seller5": {
		"product1": 76,
		"product2": 62,
		"product3": 13,
		"product4": 8
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Sellers Listings"

headings = ['Seller'] + list(data['Seller1'].keys())
ws.append(headings)

for seller in data:
	products = list(data[seller].values())
	ws.append([seller] + products)

for col in range(2, len(data['Seller1']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("SellersListings.xlsx")
